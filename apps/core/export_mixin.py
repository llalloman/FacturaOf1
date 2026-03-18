"""
Reusable export mixin for DRF ViewSets.

Adds two actions:
  GET /export_csv/   → CSV download
  GET /export_excel/ → XLSX download

Usage:
    class MyViewSet(ExportMixin, viewsets.ModelViewSet):
        export_fields = [
            ('campo_modelo', 'Encabezado Visible'),
            ('relacion__campo', 'Encabezado'),
        ]
        export_filename = 'mi_reporte'
"""

import csv
import io
from datetime import date, datetime
from decimal import Decimal

from django.http import HttpResponse
from rest_framework.decorators import action


def _resolve(obj, path):
    """Resolve a dotted/dunder path like 'comprobante__estado' on a model instance."""
    parts = path.replace('__', '.').split('.')
    val = obj
    for part in parts:
        if val is None:
            return ''
        val = getattr(val, part, None)
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, Decimal):
        return float(val)
    return val if val is not None else ''


class ExportMixin:
    """
    Mixin that adds /export_csv/ and /export_excel/ actions to a ViewSet.
    The ViewSet must define:
      - export_fields: list of (field_path, header_label)
      - export_filename: str (without extension)
    """

    export_fields: list = []
    export_filename: str = 'export'

    def _get_export_queryset(self):
        """Returns the filtered queryset for export (respects current filters)."""
        # This re-uses the ViewSet's get_queryset + filter_queryset
        qs = self.filter_queryset(self.get_queryset())
        return qs

    def _get_rows(self, qs):
        fields = self.export_fields
        headers = [h for _, h in fields]
        rows = []
        for obj in qs.iterator():
            rows.append([_resolve(obj, path) for path, _ in fields])
        return headers, rows

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        qs = self._get_export_queryset()
        headers, rows = self._get_rows(qs)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.csv"'
        response.write('\ufeff')  # BOM for Excel compatibility

        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return HttpResponse('openpyxl not installed', status=501)

        qs = self._get_export_queryset()
        headers, rows = self._get_rows(qs)

        wb = Workbook()
        ws = wb.active
        ws.title = self.export_filename[:31]

        # Header styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        thin_border = Border(
            bottom=Side(style='thin', color='D1D5DB'),
        )

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-width columns
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, min(len(rows) + 2, 102)):  # Sample first 100 rows
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Write to response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.xlsx"'
        return response
